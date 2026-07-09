"""
Conversor PDF -> XLSX (uma unica aba)
=====================================
Extrai o conteudo de um PDF preservando o maximo do layout original
(tabelas reais viram tabelas; texto livre vira linhas) e escreve TUDO
em uma unica aba do Excel, pagina apos pagina.

Adaptado do conversor_pdf.py de referencia, mas consolidado em so sheet.
"""

from __future__ import annotations

import io
import re
import traceback
from collections import Counter, defaultdict

import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TEXT_FONT = Font(size=10)
WRAP_ALIGNMENT = Alignment(wrap_text=False, vertical="top")
PAGE_SEP_FONT = Font(bold=True, italic=True, color="9CA3AF", size=9)

# Numero (pt-BR: pode ter ponto de milhar e virgula decimal) seguido de D/C
_VALOR_DIR_RE = re.compile(r"^\s*([\d.,]+)\s*([DCdc])\s*$")


class ConversorPDF:
    MIN_TABLE_ROWS = 2
    MIN_TABLE_COLS = 2

    def __init__(self, log_callback=None):
        self.log = log_callback or (lambda msg, level="info": None)

    # ------------------------------------------------------------------
    # API publica
    # ------------------------------------------------------------------
    def converter_para_excel(self, pdf_file, senha: str = "") -> dict:
        """
        Converte um PDF (caminho ou file-like) em bytes de um .xlsx
        com UMA unica aba preservando o layout.

        Retorna dict com ok, paginas, tabelas, paginas_texto, erro, dados(bytes).
        """
        resultado = {"ok": False, "paginas": 0, "tabelas": 0,
                     "paginas_texto": 0, "erro": None, "dados": b""}

        try:
            with pdfplumber.open(pdf_file, password=senha or "") as pdf:
                resultado["paginas"] = len(pdf.pages)

                wb = Workbook()
                default_sheet = wb.active
                ws = wb.create_sheet(title="PDF")

                larguras = defaultdict(int)  # coluna -> largura maxima acumulada
                linha_atual = 1

                for idx, page in enumerate(pdf.pages, start=1):
                    # Separador discreto entre paginas (exceto antes da primeira)
                    if idx > 1:
                        ws.cell(row=linha_atual, column=1,
                                value=f"— Página {idx} —").font = PAGE_SEP_FONT
                        ws.cell(row=linha_atual, column=1).alignment = WRAP_ALIGNMENT
                        linha_atual += 2  # linha do marcador + linha em branco

                    inicio_pagina = linha_atual
                    eh_tabela = self._processar_pagina(page, ws, linha_inicio=linha_atual, larguras=larguras)
                    if eh_tabela:
                        resultado["tabelas"] += 1
                    else:
                        resultado["paginas_texto"] += 1

                    # Descobre ate onde a escrita chegou
                    linha_atual = ws.max_row + 2

                    self.log(f"Página {idx}/{resultado['paginas']} processada.", level="info")

                # Remove a aba padrao vazia
                if default_sheet is not None and default_sheet.title in wb.sheetnames:
                    wb.remove(default_sheet)
                if not wb.sheetnames:
                    wb.create_sheet(title="Vazio")

                # Aplica larguras acumuladas (max por coluna)
                for col_idx, w in larguras.items():
                    ws.column_dimensions[get_column_letter(col_idx)].width = min(w + 2, 60)

                # Congela a primeira linha
                ws.freeze_panes = "A2"

                buf = io.BytesIO()
                wb.save(buf)
                resultado["dados"] = buf.getvalue()
                resultado["ok"] = True

        except Exception as e:
            nome_exc = type(e).__name__
            msg_exc = str(e).lower()
            causa = getattr(e, "__cause__", None) or getattr(e, "__context__", None)
            nome_causa = type(causa).__name__ if causa else ""
            indicios_senha = (
                "password" in nome_exc.lower() or "password" in msg_exc
                or nome_exc == "PDFPasswordIncorrect"
                or "password" in nome_causa.lower() or nome_causa == "PDFPasswordIncorrect"
            )
            if indicios_senha:
                resultado["erro"] = "PDF protegido por senha (senha incorreta ou ausente)"
            else:
                resultado["erro"] = f"{nome_exc}: {e}"
            self.log(f"Erro: {resultado['erro']}", level="error")
            self.log(traceback.format_exc(), level="debug")

        return resultado

    # ------------------------------------------------------------------
    # Processamento de pagina
    # ------------------------------------------------------------------
    def _processar_pagina(self, page, sheet, linha_inicio: int, larguras: dict) -> bool:
        texto_real = page.extract_text() or ""

        # 1. Estrutura de tabela real (linhas desenhadas) -> estrategia 'lines'
        if self._tem_estrutura_tabela_real(page):
            tabelas = self._extrair_tabelas_estrategia(page, "lines")
            validas = [t for t in tabelas
                       if t and self._qualidade_tabela_suficiente(t)
                       and not self._eh_prosa_fragmentada(t)]
            if validas:
                linha = linha_inicio
                for t in validas:
                    linha = self._escrever_tabela(sheet, t, linha, larguras)
                    linha += 1  # linha em branco entre tabelas
                return True

        # 2. Extracao por linhas de palavras (NAO parte palavras ao meio).
        #    Agrupa palavras proximas na mesma celula e so separa colunas
        #    onde existe um espaco horizontal grande de verdade.
        rows = self._extrair_linhas_palavras(page)
        if rows:
            self._escrever_tabela(sheet, rows, linha_inicio, larguras)
            return True

        # 3. Texto livre (ultima alternativa)
        self._escrever_texto(sheet, texto_real, linha_inicio)
        return False

    def _extrair_linhas_palavras(self, page):
        """
        Extrai o conteudo da pagina como linhas de celulas, agrupando
        palavras pela posicao vertical e separando colunas apenas quando
        o espaco horizontal entre palavras e grande.
        Nunca corta uma palavra ao meio.
        """
        try:
            words = page.extract_words(use_text_flow=True, keep_blank_chars=False)
        except Exception:
            return []
        if not words:
            return []

        # Agrupa palavras em linhas pela coordenada vertical (top)
        words.sort(key=lambda w: (round(w["top"] / 3), w["x0"]))
        linhas = []
        atual = []
        topo_atual = None
        for w in words:
            if topo_atual is None or abs(w["top"] - topo_atual) <= 4:
                atual.append(w)
                if topo_atual is None:
                    topo_atual = w["top"]
            else:
                linhas.append(atual)
                atual = [w]
                topo_atual = w["top"]
        if atual:
            linhas.append(atual)

        rows = []
        for linha in linhas:
            linha.sort(key=lambda w: w["x0"])
            # largura media de caractere nesta linha (para calibrar o limiar)
            chars = sum(len(w["text"]) for w in linha)
            largura_total = sum(max(0, w["x1"] - w["x0"]) for w in linha)
            char_w = (largura_total / chars) if chars else 4.0
            limiar = max(10.0, char_w * 4.0)

            celulas = []
            cel = [linha[0]]
            for prev, w in zip(linha, linha[1:]):
                gap = w["x0"] - prev["x1"]
                if gap >= limiar:  # espaco grande -> nova coluna
                    celulas.append(" ".join(c["text"] for c in cel))
                    cel = [w]
                else:
                    cel.append(w)
            celulas.append(" ".join(c["text"] for c in cel))
            rows.append(celulas)
        return rows


    def _tem_estrutura_tabela_real(self, page) -> bool:
        try:
            num_linhas = len(page.lines)
            num_rects = len(page.rects)
            num_edges = len(page.edges) if hasattr(page, "edges") else 0
        except Exception:
            return False
        return num_linhas >= 4 or num_rects >= 2 or num_edges >= 6

    def _indicio_tabela_sem_bordas(self, page, texto_real: str) -> bool:
        try:
            words = page.extract_words()
            if len(words) < 8:
                return False
            linhas = defaultdict(list)
            for w in words:
                chave = round(w["top"] / 3) * 3
                linhas[chave].append(w)
            if len(linhas) < 3:
                return False
            x_starts = Counter()
            for linha_words in linhas.values():
                for w in linha_words:
                    x_starts[round(w["x0"] / 5) * 5] += 1
            colunas_alinhadas = sum(1 for c in x_starts.values() if c >= 3)
            return colunas_alinhadas >= 2
        except Exception:
            return False

    def _eh_prosa_fragmentada(self, tabela) -> bool:
        if not tabela:
            return False
        muitas = total = 0
        for linha in tabela:
            texto = " ".join(str(c).strip() for c in linha if c and str(c).strip())
            if not texto:
                continue
            total += 1
            if len(texto.split()) > 7:
                muitas += 1
        if total == 0:
            return False
        return (muitas / total) > 0.5

    def _extrair_tabelas_estrategia(self, page, estrategia: str) -> list:
        try:
            tabelas = page.extract_tables(
                table_settings={
                    "vertical_strategy": estrategia,
                    "horizontal_strategy": estrategia,
                    "snap_tolerance": 3,
                    "join_tolerance": 3,
                    "edge_min_length": 3,
                    "text_tolerance": 1,
                }
            )
        except Exception:
            tabelas = []
        return tabelas or []

    def _parece_tabular(self, tabela) -> bool:
        if not tabela or len(tabela) < 2:
            return False
        col_counts = [len(linha) for linha in tabela]
        if not col_counts:
            return False
        contagem = Counter(col_counts)
        num_cols_mais_comum, freq = contagem.most_common(1)[0]
        consistencia = freq / len(tabela)
        if consistencia < 0.7 or num_cols_mais_comum < 2:
            return False
        comprimentos = []
        for linha in tabela:
            for celula in linha:
                if celula and str(celula).strip():
                    comprimentos.append(len(str(celula).strip()))
        if not comprimentos:
            return False
        return sum(comprimentos) / len(comprimentos) <= 60

    def _qualidade_tabela_suficiente(self, tabela) -> bool:
        if not tabela or len(tabela) < self.MIN_TABLE_ROWS:
            return False
        max_cols = max((len(linha) for linha in tabela), default=0)
        if max_cols < self.MIN_TABLE_COLS:
            return False
        return any(any(c and str(c).strip() for c in linha) for linha in tabela)

    # ------------------------------------------------------------------
    # Escrita
    # ------------------------------------------------------------------
    def _escrever_tabela(self, sheet, tabela, linha_inicio: int, larguras: dict) -> int:
        tabela = self._dividir_valor_direcao(tabela)
        linha = linha_inicio
        for i, row in enumerate(tabela):
            for j, celula in enumerate(row, start=1):
                valor = self._limpar_celula(celula)
                c = sheet.cell(row=linha, column=j, value=valor)
                c.font = TEXT_FONT
                c.alignment = WRAP_ALIGNMENT
                if i == 0 and linha_inicio == 1:
                    c.font = HEADER_FONT
                    c.fill = HEADER_FILL
            linha += 1
        self._atualizar_larguras(tabela, larguras)
        return linha

    def _escrever_texto(self, sheet, texto: str, linha_inicio: int) -> int:
        if not texto.strip():
            c = sheet.cell(row=linha_inicio, column=1, value="[Página sem texto extraído]")
            c.font = Font(italic=True, color="999999")
            return linha_inicio + 1
        linhas = texto.split("\n")
        for i, linha in enumerate(linhas):
            c = sheet.cell(row=linha_inicio + i, column=1, value=linha)
            c.font = TEXT_FONT
            c.alignment = WRAP_ALIGNMENT
        return linha_inicio + len(linhas)

    # ------------------------------------------------------------------
    # Utilitarios
    # ------------------------------------------------------------------
    def _dividir_valor_direcao(self, tabela):
        """Separa celulas do tipo '72,12D' em duas: '72,12' e 'D'."""
        nova = []
        for linha in tabela:
            nova_linha = []
            for celula in linha:
                txt = self._limpar_celula(celula)
                m = _VALOR_DIR_RE.match(txt)
                if m:
                    nova_linha.append(m.group(1))
                    nova_linha.append(m.group(2).upper())
                else:
                    nova_linha.append(celula)
            nova.append(nova_linha)
        return nova

    @staticmethod
    def _limpar_celula(valor) -> str:
        if valor is None:
            return ""
        texto = str(valor).strip()
        texto = texto.replace("\r\n", " ").replace("\n", " ").replace("\r", " ")
        while "  " in texto:
            texto = texto.replace("  ", " ")
        return texto

    @staticmethod
    def _atualizar_larguras(tabela, larguras: dict) -> None:
        if not tabela:
            return
        num_cols = max((len(linha) for linha in tabela), default=0)
        for col_idx in range(1, num_cols + 1):
            max_len = 10
            for linha in tabela:
                if col_idx <= len(linha):
                    celula = linha[col_idx - 1]
                    if celula:
                        for sub in str(celula).split("\n"):
                            max_len = max(max_len, len(sub))
            larguras[col_idx] = max(larguras.get(col_idx, 0), max_len)
