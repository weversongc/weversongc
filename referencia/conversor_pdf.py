#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Conversor PDF -> XLSX
=====================
Aplicativo desktop com interface grafica que converte arquivos PDF em arquivos
Excel (.xlsx), um .xlsx para cada PDF (sem mesclar).

Recursos:
  - Interface grafica (Tkinter) com seletor de multiplos PDFs
  - Deteccao automatica por pagina: tabelas viram tabelas no Excel;
    texto livre vira uma linha por celula
  - Uma aba (sheet) por pagina do PDF dentro de cada .xlsx
  - Tratamento de PDF protegido por senha
  - Log de erros (arquivo .log ao lado dos PDFs)
  - Pergunta sobrescrita quando o .xlsx ja existe
  - Barra de progresso com nome do arquivo atual

Como executar (desenvolvimento):
    python conversor_pdf.py

Como gerar o .exe (Windows):
    build.bat
"""

from __future__ import annotations

import os
import sys
import traceback
import threading
import datetime
from pathlib import Path
from typing import List, Optional, Tuple

# Dependencias externas
import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# GUI
import tkinter as tk
from tkinter import ttk, filedialog, messagebox

# ---------------------------------------------------------------------------
# Configuracoes gerais
# ---------------------------------------------------------------------------
APP_TITLE = "Conversor PDF -> XLSX"
APP_VERSION = "1.0.0"

# Estilos visuais
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TEXT_FONT = Font(size=10)
WRAP_ALIGNMENT = Alignment(wrap_text=True, vertical="top")


# ---------------------------------------------------------------------------
# Nucleo de conversao (sem GUI) - facilita testes automatizados
# ---------------------------------------------------------------------------
class ConversionError(Exception):
    """Erro de conversao com classificacao para o log."""


class ConversorPDF:
    """Converte um unico PDF em um unico XLSX."""

    # Limiares para decidir se uma pagina e "tabela" ou "texto"
    MIN_TABLE_ROWS = 2
    MIN_TABLE_COLS = 2

    def __init__(self, log_callback=None):
        self.log = log_callback or (lambda msg, level="info": None)

    # ------------------------------------------------------------------
    # API publica
    # ------------------------------------------------------------------
    def converter_pdf_para_xlsx(
        self,
        pdf_path: str,
        xlsx_path: str,
        senha: Optional[str] = None,
    ) -> dict:
        """
        Converte um PDF em XLSX.

        Retorna um dict:
            {
                "ok": bool,
                "paginas": int,
                "tabelas": int,
                "paginas_texto": int,
                "erro": str | None,
            }
        """
        resultado = {
            "ok": False,
            "paginas": 0,
            "tabelas": 0,
            "paginas_texto": 0,
            "erro": None,
        }

        if not os.path.isfile(pdf_path):
            resultado["erro"] = f"Arquivo PDF nao encontrado: {pdf_path}"
            return resultado

        # Verificar se e protegido antes de abrir (mensagem de erro mais clara)
        if not senha:
            senha_usar = ""
        else:
            senha_usar = senha

        try:
            with pdfplumber.open(pdf_path, password=senha_usar) as pdf:
                resultado["paginas"] = len(pdf.pages)
                wb = Workbook()
                # Remover a sheet padrao criada pelo openpyxl
                default_sheet = wb.active

                for idx, page in enumerate(pdf.pages, start=1):
                    sheet_name = f"Pagina_{idx}"
                    sheet = wb.create_sheet(title=sheet_name)

                    try:
                        tem_tabela = self._processar_pagina(page, sheet)
                        if tem_tabela:
                            resultado["tabelas"] += 1
                        else:
                            resultado["paginas_texto"] += 1
                    except Exception as e:
                        # Erro em uma pagina nao deve parar todo o PDF
                        self.log(
                            f"  Aviso: pagina {idx} do PDF "
                            f"'{os.path.basename(pdf_path)}' gerou erro: {e}",
                            level="warn",
                        )
                        sheet["A1"] = f"[Erro ao extrair pagina {idx}] {e}"

                # Remover a sheet padrao se ainda existir e estiver vazia
                if default_sheet is not None and default_sheet.title in wb.sheetnames:
                    wb.remove(default_sheet)

                # Garantir que existe ao menos uma sheet
                if not wb.sheetnames:
                    wb.create_sheet(title="Vazio")

                wb.save(xlsx_path)
                resultado["ok"] = True

        except Exception as e:
            nome_exc = type(e).__name__
            msg_exc = str(e).lower()
            # Inspecionar causa interna (PdfminerException wrapa PDFPasswordIncorrect)
            causa = getattr(e, "__cause__", None) or getattr(e, "__context__", None)
            nome_causa = type(causa).__name__ if causa else ""
            # Detectar erro de senha em diferentes versoes do pdfplumber/pdfminer
            indicios_senha = (
                "password" in nome_exc.lower()
                or "password" in msg_exc
                or nome_exc == "PDFPasswordIncorrect"
                or "password" in nome_causa.lower()
                or nome_causa == "PDFPasswordIncorrect"
            )
            if indicios_senha:
                resultado["erro"] = "PDF protegido por senha (senha incorreta ou ausente)"
            else:
                resultado["erro"] = f"{nome_exc}: {e}"
            self.log(
                f"Erro ao converter '{os.path.basename(pdf_path)}': {resultado['erro']}",
                level="error",
            )
            self.log(traceback.format_exc(), level="debug")

        return resultado

    # ------------------------------------------------------------------
    # Processamento de pagina
    # ------------------------------------------------------------------
    def _processar_pagina(self, page, sheet) -> bool:
        """
        Processa uma pagina do PDF e escreve na sheet do Excel.

        Retorna True se a pagina foi identificada como tabela,
        False se foi tratada como texto livre.

        Estrategia:
          1. Verifica se a pagina tem linhas/retangulos desenhados (sinal de tabela real).
          2. Se sim, tenta extrair tabelas com estrategia 'lines'.
          3. Valida se a tabela extraida nao e na verdade prosa fragmentada.
          4. Se a tabela for confiavel, usa ela.
          5. Caso contrario, extrai como texto livre.
        """
        # Sempre obter o texto real da pagina
        texto_real = page.extract_text() or ""

        # 1. Verificar se a pagina tem estrutura de tabela real (linhas desenhadas)
        tem_estrutura_tabela = self._tem_estrutura_tabela_real(page)

        if tem_estrutura_tabela:
            tabelas = self._extrair_tabelas_estrategia(page, "lines")
            tabelas_validas = [
                t for t in tabelas
                if t and self._qualidade_tabela_suficiente(t)
                and not self._eh_prosa_fragmentada(t)
            ]
            if tabelas_validas:
                linha_atual = 1
                for tabela in tabelas_validas:
                    linha_atual = self._escrever_tabela(
                        sheet, tabela, linha_inicio=linha_atual
                    )
                    linha_atual += 1  # linha em branco entre tabelas
                return True

        # 2. Tentar estrategia por texto apenas se houver indicios de tabela
        #    (evita falsos positivos em paginas puramente textuais)
        if tem_estrutura_tabela or self._indicio_tabela_sem_bordas(page, texto_real):
            tabelas_texto = self._extrair_tabelas_estrategia(page, "text")
            tabelas_validas = [
                t for t in tabelas_texto
                if t and self._qualidade_tabela_suficiente(t)
                and self._parece_tabular(t)
                and not self._eh_prosa_fragmentada(t)
            ]
            if tabelas_validas:
                linha_atual = 1
                for tabela in tabelas_validas:
                    linha_atual = self._escrever_tabela(
                        sheet, tabela, linha_inicio=linha_atual
                    )
                    linha_atual += 1
                return True

        # 3. Texto livre (prosa)
        self._escrever_texto(sheet, texto_real)
        return False

    def _tem_estrutura_tabela_real(self, page) -> bool:
        """
        Verifica se a pagina tem linhas/retangulos desenhados que sugerem
        uma tabela real (com bordas).
        """
        try:
            num_linhas = len(page.lines)
            num_rects = len(page.rects)
            num_edges = len(page.edges) if hasattr(page, "edges") else 0
        except Exception:
            return False
        # Heuristica: pelo menos 4 linhas OU 2 retangulos OU 6 edges
        return num_linhas >= 4 or num_rects >= 2 or num_edges >= 6

    def _indicio_tabela_sem_bordas(self, page, texto_real: str) -> bool:
        """
        Detecta indicios de tabela sem bordas, baseado em alinhamento de
        colunas no texto. Retorna True se houver colunas visivelmente alinhadas.
        """
        try:
            words = page.extract_words()
            if len(words) < 8:
                return False
            # Agrupar por linha (top arredondado)
            from collections import defaultdict
            linhas = defaultdict(list)
            for w in words:
                # Agrupar por posicao vertical (top), com tolerancia de 3px
                chave = round(w["top"] / 3) * 3
                linhas[chave].append(w)
            if len(linhas) < 3:
                return False
            # Contar quantas linhas tem 3+ palavras comecando em posicoes X similares
            # (sinal de colunas alinhadas)
            from collections import Counter
            x_starts = Counter()
            for linha_words in linhas.values():
                for w in linha_words:
                    x_starts[round(w["x0"] / 5) * 5] += 1
            # Se ha 2+ posicoes X com 3+ palavras cada, parece tabela
            colunas_alinhadas = sum(1 for c in x_starts.values() if c >= 3)
            return colunas_alinhadas >= 2
        except Exception:
            return False

    def _eh_prosa_fragmentada(self, tabela) -> bool:
        """
        Detecta se uma 'tabela' extraida e na verdade prosa fragmentada em colunas.

        Sinais de prosa:
          - Ao juntar celulas de cada linha, o resultado tem MUITAS palavras
            (tabelas reais tem poucas palavras por linha)
          - Celulas terminam no meio de palavras (sem pontuacao/espaco no final)
        """
        if not tabela:
            return False

        linhas_com_muitas_palavras = 0
        total_linhas = 0
        for linha in tabela:
            texto = " ".join(
                str(c).strip() for c in linha if c and str(c).strip()
            )
            if not texto:
                continue
            total_linhas += 1
            num_palavras = len(texto.split())
            # Tabela real: tipicamente 1-6 palavras por linha
            # Prosa: tipicamente 8+ palavras por linha
            if num_palavras > 7:
                linhas_com_muitas_palavras += 1

        if total_linhas == 0:
            return False
        return (linhas_com_muitas_palavras / total_linhas) > 0.5

    def _extrair_tabelas_estrategia(self, page, estrategia: str) -> list:
        """Extrai tabelas usando a estrategia especificada ('lines' ou 'text')."""
        try:
            tabelas = page.extract_tables(
                table_settings={
                    "vertical_strategy": estrategia,
                    "horizontal_strategy": estrategia,
                    "snap_tolerance": 3,
                    "join_tolerance": 3,
                    "edge_min_length": 3,
                    "text_tolerance": 1,
                }
            )
        except Exception:
            tabelas = []
        return tabelas or []

    def _parece_tabular(self, tabela) -> bool:
        """
        Heuristica para distinguir tabela real de prosa fragmentada.

        Criterios para considerar tabular:
          - Linhas tem numero de colunas consistente (pelo menos 70% iguais)
          - Celulas sao curtas (media < 60 chars) - prosa tem celulas longas
          - Pelo menos 2 colunas tem conteudo em mais de uma linha
        """
        if not tabela or len(tabela) < 2:
            return False

        # Contar distribuicao de colunas por linha
        col_counts = [len(linha) for linha in tabela]
        if not col_counts:
            return False

        # Moda do numero de colunas
        from collections import Counter
        contagem = Counter(col_counts)
        num_cols_mais_comum, freq = contagem.most_common(1)[0]
        consistencia = freq / len(tabela)
        if consistencia < 0.7:
            return False
        if num_cols_mais_comum < 2:
            return False

        # Comprimento medio das celulas nao vazias
        comprimentos = []
        for linha in tabela:
            for celula in linha:
                if celula and str(celula).strip():
                    comprimentos.append(len(str(celula).strip()))
        if not comprimentos:
            return False
        media_compr = sum(comprimentos) / len(comprimentos)
        if media_compr > 60:
            return False

        return True

    def _qualidade_tabela_suficiente(self, tabela) -> bool:
        """Verifica se a tabela extraida tem qualidade suficiente para usar."""
        if not tabela:
            return False
        if len(tabela) < self.MIN_TABLE_ROWS:
            return False
        # Contar colunas na linha mais larga
        max_cols = max((len(linha) for linha in tabela), default=0)
        if max_cols < self.MIN_TABLE_COLS:
            return False
        # Pelo menos uma celula deve ter conteudo
        tem_conteudo = any(
            any(celula and str(celula).strip() for celula in linha)
            for linha in tabela
        )
        return tem_conteudo

    def _escrever_tabela(
        self, sheet, tabela, linha_inicio: int = 1
    ) -> int:
        """
        Escreve uma tabela na sheet a partir de linha_inicio.
        Retorna a proxima linha disponivel.
        """
        linha = linha_inicio
        for i, row in enumerate(tabela):
            for j, celula in enumerate(row, start=1):
                valor = self._limpar_celula(celula)
                celula_excel = sheet.cell(row=linha, column=j, value=valor)
                celula_excel.font = TEXT_FONT
                celula_excel.alignment = WRAP_ALIGNMENT
                # Primeira linha como cabecalho estilizado
                if i == 0 and linha_inicio == 1:
                    celula_excel.font = HEADER_FONT
                    celula_excel.fill = HEADER_FILL
            linha += 1

        # Auto-ajustar largura das colunas (aproximado)
        self._ajustar_larguras(sheet, tabela, linha_inicio)
        return linha

    def _escrever_texto(self, sheet, texto: str) -> None:
        """Escreve texto livre, uma linha do PDF por celula na coluna A."""
        if not texto.strip():
            sheet["A1"] = "[Pagina sem texto extraido]"
            sheet["A1"].font = Font(italic=True, color="999999")
            return

        linhas = texto.split("\n")
        for i, linha in enumerate(linhas, start=1):
            celula = sheet.cell(row=i, column=1, value=linha)
            celula.font = TEXT_FONT
            celula.alignment = WRAP_ALIGNMENT

        # Largura confortavel para leitura de texto
        sheet.column_dimensions["A"].width = 120

    # ------------------------------------------------------------------
    # Utilitarios
    # ------------------------------------------------------------------
    @staticmethod
    def _limpar_celula(valor) -> str:
        """Limpa e normaliza o conteudo de uma celula extraida da tabela."""
        if valor is None:
            return ""
        texto = str(valor).strip()
        # Remover quebras de linha internas que atrapalham o Excel
        texto = texto.replace("\r\n", " ").replace("\n", " ").replace("\r", " ")
        # Colapsar espacos multiplus
        while "  " in texto:
            texto = texto.replace("  ", " ")
        return texto

    @staticmethod
    def _ajustar_larguras(sheet, tabela, linha_inicio: int) -> None:
        """Ajusta largura das colunas baseado no conteudo."""
        if not tabela:
            return
        num_cols = max((len(linha) for linha in tabela), default=0)
        for col_idx in range(1, num_cols + 1):
            max_len = 10  # minimo
            for linha in tabela:
                if col_idx <= len(linha):
                    celula = linha[col_idx - 1]
                    if celula:
                        # Comprimento aproximado da maior linha
                        for sub in str(celula).split("\n"):
                            max_len = max(max_len, len(sub))
            # Limite para nao criar colunas absurdas
            max_len = min(max_len + 2, 60)
            col_letter = get_column_letter(col_idx)
            sheet.column_dimensions[col_letter].width = max_len


# ---------------------------------------------------------------------------
# Interface Grafica (Tkinter)
# ---------------------------------------------------------------------------
class ConversorApp:
    def __init__(self, root: tk.Tk):
        self.root = root
        self.root.title(f"{APP_TITLE} v{APP_VERSION}")
        self.root.geometry("780x560")
        self.root.minsize(640, 460)

        # Estado
        self.arquivos_selecionados: List[str] = []
        self.conversor = ConversorPDF(log_callback=self._log_interno)
        self.processando = False

        self._construir_ui()

    # ------------------------------------------------------------------
    # Construcao da UI
    # ------------------------------------------------------------------
    def _construir_ui(self):
        # Estilo
        style = ttk.Style()
        try:
            style.theme_use("clam")
        except tk.TclError:
            pass

        # Container principal
        main = ttk.Frame(self.root, padding=12)
        main.pack(fill=tk.BOTH, expand=True)

        # Titulo
        titulo = ttk.Label(
            main,
            text=f"{APP_TITLE}",
            font=("Segoe UI", 16, "bold"),
        )
        titulo.pack(anchor=tk.W, pady=(0, 4))

        subtitulo = ttk.Label(
            main,
            text="Selecione um ou mais PDFs. Cada PDF sera convertido em um arquivo .xlsx separado.",
            font=("Segoe UI", 9),
            foreground="#555555",
        )
        subtitulo.pack(anchor=tk.W, pady=(0, 12))

        # Botoes de acao
        botoes = ttk.Frame(main)
        botoes.pack(fill=tk.X, pady=(0, 8))

        self.btn_selecionar = ttk.Button(
            botoes, text="Selecionar PDFs...", command=self.selecionar_pdfs
        )
        self.btn_selecionar.pack(side=tk.LEFT, padx=(0, 6))

        self.btn_converter = ttk.Button(
            botoes,
            text="Converter para Excel",
            command=self.iniciar_conversao,
            state=tk.DISABLED,
        )
        self.btn_converter.pack(side=tk.LEFT, padx=(0, 6))

        self.btn_limpar = ttk.Button(
            botoes, text="Limpar lista", command=self.limpar_lista, state=tk.DISABLED
        )
        self.btn_limpar.pack(side=tk.LEFT)

        # Lista de arquivos
        lista_frame = ttk.LabelFrame(main, text="Arquivos selecionados", padding=8)
        lista_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 8))

        self.lista = tk.Listbox(lista_frame, selectmode=tk.EXTENDED, font=("Consolas", 9))
        self.lista.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        scrollbar = ttk.Scrollbar(lista_frame, orient=tk.VERTICAL, command=self.lista.yview)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.lista.config(yscrollcommand=scrollbar.set)

        # Progresso
        progresso_frame = ttk.LabelFrame(main, text="Progresso", padding=8)
        progresso_frame.pack(fill=tk.X, pady=(0, 8))

        self.lbl_status = ttk.Label(progresso_frame, text="Pronto.", font=("Segoe UI", 9))
        self.lbl_status.pack(anchor=tk.W, pady=(0, 4))

        self.progress = ttk.Progressbar(
            progresso_frame, orient=tk.HORIZONTAL, length=100, mode="determinate"
        )
        self.progress.pack(fill=tk.X)

        self.lbl_percent = ttk.Label(progresso_frame, text="0%", font=("Segoe UI", 9))
        self.lbl_percent.pack(anchor=tk.E, pady=(2, 0))

        # Log
        log_frame = ttk.LabelFrame(main, text="Log", padding=8)
        log_frame.pack(fill=tk.BOTH, expand=True)

        self.log_text = tk.Text(
            log_frame, height=8, font=("Consolas", 9), state=tk.DISABLED, wrap=tk.WORD
        )
        self.log_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        log_scroll = ttk.Scrollbar(log_frame, orient=tk.VERTICAL, command=self.log_text.yview)
        log_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        self.log_text.config(yscrollcommand=log_scroll.set)

    # ------------------------------------------------------------------
    # Acoes
    # ------------------------------------------------------------------
    def selecionar_pdfs(self):
        arquivos = filedialog.askopenfilenames(
            title="Selecionar arquivos PDF",
            filetypes=[("Arquivos PDF", "*.pdf"), ("Todos os arquivos", "*.*")],
        )
        if not arquivos:
            return

        self.arquivos_selecionados = list(arquivos)
        self.lista.delete(0, tk.END)
        for arq in self.arquivos_selecionados:
            self.lista.insert(tk.END, os.path.basename(arq))

        self._log(f"{len(self.arquivos_selecionados)} arquivo(s) selecionado(s).")
        self.btn_converter.config(state=tk.NORMAL)
        self.btn_limpar.config(state=tk.NORMAL)
        self.lbl_status.config(text=f"{len(self.arquivos_selecionados)} PDF(s) pronto(s) para conversao.")

    def limpar_lista(self):
        self.arquivos_selecionados.clear()
        self.lista.delete(0, tk.END)
        self.btn_converter.config(state=tk.DISABLED)
        self.btn_limpar.config(state=tk.DISABLED)
        self.lbl_status.config(text="Pronto.")
        self._atualizar_progresso(0, 1)

    def iniciar_conversao(self):
        if not self.arquivos_selecionados:
            messagebox.showwarning(APP_TITLE, "Selecione pelo menos um arquivo PDF.")
            return
        if self.processando:
            return

        # Confirmar sobrescrita previa (configuracao)
        self.processando = True
        self._alternar_botoes(enabled=False)
        self._log("=" * 60)
        self._log(f"Iniciando conversao de {len(self.arquivos_selecionados)} arquivo(s).")

        # Rodar em thread separada para nao travar a UI
        thread = threading.Thread(target=self._processar_lote, daemon=True)
        thread.start()

    # ------------------------------------------------------------------
    # Processamento em lote (em thread)
    # ------------------------------------------------------------------
    def _processar_lote(self):
        arquivos = list(self.arquivos_selecionados)
        total = len(arquivos)
        sucessos = 0
        falhas = []
        entradas_log = []  # (arquivo, status, mensagem)

        for i, pdf_path in enumerate(arquivos, start=1):
            nome = os.path.basename(pdf_path)
            self._status(f"Processando ({i}/{total}): {nome}")
            self._log(f"[{i}/{total}] {nome}")

            xlsx_path = os.path.splitext(pdf_path)[0] + ".xlsx"

            # Verificar sobrescrita
            if os.path.exists(xlsx_path):
                acao = self._perguntar_sobrescrita(nome)
                if acao == "pular":
                    self._log("  -> Pulado pelo usuario (arquivo .xlsx ja existe).")
                    entradas_log.append((pdf_path, "pulou", "Arquivo .xlsx ja existia"))
                    self._atualizar_progresso(i, total)
                    continue
                # acao == "sobrescrever" -> prosseguir

            # Verificar se PDF e protegido
            senha = None
            if self._pdf_protegido(pdf_path):
                senha = self._pedir_senha(nome)
                if senha is None:
                    self._log("  -> Pulado (usuario cancelou senha).")
                    entradas_log.append((pdf_path, "pulou", "Senha nao fornecida"))
                    self._atualizar_progresso(i, total)
                    continue

            # Converter
            resultado = self.conversor.converter_pdf_para_xlsx(
                pdf_path, xlsx_path, senha=senha
            )

            if resultado["ok"]:
                sucessos += 1
                msg = (
                    f"  OK -> {os.path.basename(xlsx_path)} "
                    f"({resultado['paginas']} pag, "
                    f"{resultado['tabelas']} tab, "
                    f"{resultado['paginas_texto']} texto)"
                )
                self._log(msg)
                entradas_log.append((pdf_path, "ok", msg.strip()))
            else:
                falhas.append((nome, resultado["erro"]))
                self._log(f"  FALHA: {resultado['erro']}", level="error")
                entradas_log.append((pdf_path, "falha", resultado["erro"]))

            self._atualizar_progresso(i, total)

        # Gerar arquivo de log
        log_path = self._gerar_log_arquivo(entradas_log, arquivos[0] if arquivos else None)

        # Resumo
        self._log("=" * 60)
        self._log(f"Concluido. Sucessos: {sucessos} | Falhas: {len(falhas)} | Total: {total}")
        if log_path:
            self._log(f"Log salvo em: {log_path}")
        if falhas:
            self._log("Arquivos com falha:")
            for nome, erro in falhas:
                self._log(f"  - {nome}: {erro}")

        self._status(
            f"Concluido: {sucessos} sucesso(s), {len(falhas)} falha(s)."
        )
        self.processando = False
        self._alternar_botoes(enabled=True)

        # Aviso final
        if falhas:
            messagebox.showwarning(
                APP_TITLE,
                f"Conversao concluida com {len(falhas)} falha(s) de {total}.\n"
                f"Consulte o log para detalhes.",
            )
        else:
            messagebox.showinfo(
                APP_TITLE,
                f"Conversao concluida com sucesso!\n{sucessos} arquivo(s) .xlsx gerado(s).",
            )

    # ------------------------------------------------------------------
    # Helpers de UI / interacao
    # ------------------------------------------------------------------
    def _pdf_protegido(self, pdf_path: str) -> bool:
        """Verifica se o PDF exige senha para abertura."""
        try:
            with pdfplumber.open(pdf_path, password="") as pdf:
                # Acessar primeira pagina para forcar a leitura
                if pdf.pages:
                    _ = pdf.pages[0].extract_text()
            return False
        except Exception as e:
            nome_exc = type(e).__name__
            msg_exc = str(e).lower()
            # Inspecionar causa interna
            causa = getattr(e, "__cause__", None) or getattr(e, "__context__", None)
            nome_causa = type(causa).__name__ if causa else ""
            # Se for erro de senha, o PDF e protegido
            if (
                "password" in nome_exc.lower()
                or "password" in msg_exc
                or nome_exc == "PDFPasswordIncorrect"
                or "password" in nome_causa.lower()
                or nome_causa == "PDFPasswordIncorrect"
            ):
                return True
            # Outros erros (PDF corrompido, etc.) nao sao de senha
            # Deixar o processo de conversao lidar com eles depois
            self._log(
                f"Aviso ao verificar protecao de '{os.path.basename(pdf_path)}': {e}",
                level="warn",
            )
            return False

    def _pedir_senha(self, nome: str) -> Optional[str]:
        """Abre dialogo pedindo senha. Retorna None se cancelado."""
        dialog = tk.Toplevel(self.root)
        dialog.title("PDF protegido")
        dialog.geometry("420x180")
        dialog.transient(self.root)
        dialog.grab_set()

        ttk.Label(
            dialog,
            text=f"O arquivo abaixo esta protegido por senha:\n{nome}",
            font=("Segoe UI", 9),
            wraplength=380,
        ).pack(padx=12, pady=(12, 8), anchor=tk.W)

        senha_var = tk.StringVar()
        entry = ttk.Entry(dialog, textvariable=senha_var, show="*")
        entry.pack(padx=12, pady=4, fill=tk.X)
        entry.focus_set()

        resultado = {"senha": None}

        def confirmar():
            resultado["senha"] = senha_var.get() or None
            dialog.destroy()

        def cancelar():
            resultado["senha"] = None
            dialog.destroy()

        btns = ttk.Frame(dialog)
        btns.pack(pady=8)
        ttk.Button(btns, text="Cancelar", command=cancelar).pack(side=tk.LEFT, padx=4)
        ttk.Button(btns, text="Confirmar", command=confirmar).pack(side=tk.LEFT, padx=4)

        dialog.bind("<Return>", lambda e: confirmar())
        dialog.bind("<Escape>", lambda e: cancelar())

        self.root.wait_window(dialog)
        return resultado["senha"]

    def _perguntar_sobrescrita(self, nome: str) -> str:
        """Pergunta se usuario quer sobrescrever. Retorna 'sobrescrever' ou 'pular'."""
        # Roda na thread principal via after
        resultado = {"acao": "pular"}

        def perguntar():
            resp = messagebox.askyesno(
                APP_TITLE,
                f"O arquivo de saida ja existe:\n{nome}.xlsx\n\nDeseja sobrescrever?",
            )
            resultado["acao"] = "sobrescrever" if resp else "pular"

        self.root.after(0, perguntar)
        # Aguardar (nao ideal, mas simples)
        self.root.after(100, lambda: None)
        # Como estamos em thread, usar askyesno diretamente funciona no Windows/Linux
        try:
            resp = messagebox.askyesno(
                APP_TITLE,
                f"O arquivo de saida ja existe:\n{nome}.xlsx\n\nDeseja sobrescrever?",
            )
            return "sobrescrever" if resp else "pular"
        except Exception:
            return resultado["acao"]

    def _gerar_log_arquivo(self, entradas, pdf_referencia: Optional[str]) -> Optional[str]:
        if not entradas:
            return None
        # Salvar o log na mesma pasta do primeiro PDF
        pasta = (
            os.path.dirname(pdf_referencia)
            if pdf_referencia
            else os.getcwd()
        )
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        log_path = os.path.join(pasta, f"conversao_pdf_xlsx_{timestamp}.log")
        try:
            with open(log_path, "w", encoding="utf-8") as f:
                f.write(f"Log de conversao PDF -> XLSX\n")
                f.write(f"Data: {datetime.datetime.now().isoformat()}\n")
                f.write(f"Aplicativo: {APP_TITLE} v{APP_VERSION}\n")
                f.write("=" * 60 + "\n\n")
                for arq, status, msg in entradas:
                    f.write(f"Arquivo: {arq}\n")
                    f.write(f"Status:  {status.upper()}\n")
                    f.write(f"Detalhe: {msg}\n")
                    f.write("-" * 60 + "\n")
        except Exception as e:
            self._log(f"Nao foi possivel salvar o arquivo de log: {e}", level="error")
            return None
        return log_path

    # ------------------------------------------------------------------
    # Atualizadores de UI (thread-safe via after)
    # ------------------------------------------------------------------
    def _status(self, msg: str):
        self.root.after(0, lambda: self.lbl_status.config(text=msg))

    def _atualizar_progresso(self, atual: int, total: int):
        def _update():
            self.progress["maximum"] = total
            self.progress["value"] = atual
            pct = int((atual / total) * 100) if total else 0
            self.lbl_percent.config(text=f"{pct}%")
        self.root.after(0, _update)

    def _log(self, msg: str, level: str = "info"):
        self.root.after(0, lambda: self._log_ui(msg, level))

    def _log_interno(self, msg: str, level: str = "info"):
        # Callback usado pelo ConversorPDF
        self._log(msg, level)

    def _log_ui(self, msg: str, level: str = "info"):
        cores = {
            "info": "#222222",
            "warn": "#AA6600",
            "error": "#CC0000",
            "debug": "#888888",
        }
        self.log_text.config(state=tk.NORMAL)
        self.log_text.insert(tk.END, msg + "\n")
        # Colorir a ultima linha
        cor = cores.get(level, "#222222")
        ultima_linha = int(self.log_text.index("end-1c").split(".")[0]) - 1
        if ultima_linha > 0:
            inicio = f"{ultima_linha}.0"
            fim = f"{ultima_linha}.end"
            self.log_text.tag_add(cor, inicio, fim)
            self.log_text.tag_config(cor, foreground=cor)
        self.log_text.see(tk.END)
        self.log_text.config(state=tk.DISABLED)

    def _alternar_botoes(self, enabled: bool):
        estado = tk.NORMAL if enabled else tk.DISABLED
        self.root.after(0, lambda: self.btn_selecionar.config(state=estado))
        self.root.after(0, lambda: self.btn_limpar.config(state=estado))
        # Manter "Converter" desabilitado se nao houver arquivos
        if enabled and self.arquivos_selecionados:
            self.root.after(0, lambda: self.btn_converter.config(state=tk.NORMAL))
        else:
            self.root.after(0, lambda: self.btn_converter.config(state=tk.DISABLED))


# ---------------------------------------------------------------------------
# Modo CLI (para testes automatizados)
# ---------------------------------------------------------------------------
def cli_converter(arquivos: List[str], senha: Optional[str] = None) -> int:
    """Converte arquivos via linha de comando. Retorna numero de falhas."""
    conversor = ConversorPDF(log_callback=lambda m, l="info": print(m))
    falhas = 0
    for pdf_path in arquivos:
        xlsx_path = os.path.splitext(pdf_path)[0] + ".xlsx"
        print(f"Convertendo: {pdf_path}")
        resultado = conversor.converter_pdf_para_xlsx(pdf_path, xlsx_path, senha=senha)
        if resultado["ok"]:
            print(
                f"  OK: {resultado['paginas']} pag, "
                f"{resultado['tabelas']} tab, "
                f"{resultado['paginas_texto']} texto"
            )
        else:
            print(f"  FALHA: {resultado['erro']}")
            falhas += 1
    return falhas


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------
def main():
    # Se houver argumentos, rodar em modo CLI
    if len(sys.argv) > 1:
        arquivos = [a for a in sys.argv[1:] if a.lower().endswith(".pdf")]
        if arquivos:
            falhas = cli_converter(arquivos)
            sys.exit(0 if falhas == 0 else 1)

    # Modo GUI
    root = tk.Tk()
    app = ConversorApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
