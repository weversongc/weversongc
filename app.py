import streamlit as st
import pandas as pd
import pdfplumber
import io
from supabase import create_client, Client
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment

# =========================================================================
# 1. CONFIGURAÇÃO DE PÁGINA
# =========================================================================
st.set_page_config(page_title="Weversongc - Automações", layout="wide")


# =========================================================================
# 2. CONFIGURAÇÃO DE CREDENCIAIS (lidas de st.secrets)
#    No Streamlit Cloud: Settings -> Secrets
#    Local: arquivo .streamlit/secrets.toml
# =========================================================================
SUPABASE_URL = st.secrets.get("SUPABASE_URL", "")
SUPABASE_KEY = st.secrets.get("SUPABASE_KEY", "")
NOME_BUCKET = st.secrets.get("NOME_BUCKET", "arquivo-usuários")


@st.cache_resource
def iniciar_supabase() -> Client | None:
    if not SUPABASE_URL or not SUPABASE_KEY:
        return None
    try:
        return create_client(SUPABASE_URL, SUPABASE_KEY)
    except Exception:
        return None


supabase = iniciar_supabase()


# =========================================================================
# 3. FUNÇÕES AUXILIARES
# =========================================================================
def enviar_para_supabase(nome_arquivo: str, dados_bytes: bytes) -> None:
    if supabase is None:
        st.info("ℹ️ Supabase não configurado. Arquivo disponível apenas para download.")
        return
    try:
        supabase.storage.from_(NOME_BUCKET).upload(
            path=nome_arquivo,
            file=dados_bytes,
            file_options={
                "content-type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            },
        )
        st.success(f"☁️ Arquivo '{nome_arquivo}' salvo no Supabase Storage!")
    except Exception as e:
        st.warning(f"Aviso ao salvar na nuvem: {e}")


# =========================================================================
# 4. NAVEGAÇÃO LATERAL
# =========================================================================
st.sidebar.title("🤖 Painel de Controle")
st.sidebar.markdown("**Weversongc**")

opcao = st.sidebar.radio(
    "Selecione a ferramenta:",
    [
        "📄 Conversor de PDF para Excel",
        "📊 Conciliação Bancária",
        "💸 Diferenças Braslog vs MH",
    ],
)

st.sidebar.markdown("---")
if supabase is not None:
    st.sidebar.caption("☁️ Conectado ao Supabase")
else:
    st.sidebar.caption("🔒 Supabase offline (modo local)")


# =========================================================================
# MÓDULO: CONVERSOR DE PDF
# =========================================================================
if opcao == "📄 Conversor de PDF para Excel":
    st.title("📄 Conversor Inteligente de PDF para XLSX")
    st.markdown("Converta seus relatórios em PDF para abas estruturadas no Excel de forma automática.")

    arquivo_pdf = st.file_uploader("Arraste seu arquivo PDF aqui", type=["pdf"])

    if arquivo_pdf is not None:
        if st.button("🚀 Iniciar Conversão"):
            with st.spinner("Processando páginas do PDF..."):
                wb = Workbook()
                if "Sheet" in wb.sheetnames:
                    wb.remove(wb["Sheet"])

                with pdfplumber.open(arquivo_pdf) as pdf:
                    for i, pagina in enumerate(pdf.pages):
                        nome_aba = f"Pagina_{i+1}"
                        ws = wb.create_sheet(title=nome_aba)

                        tabelas = pagina.extract_tables()
                        if tabelas:
                            for tabela in tabelas:
                                for linha in tabela:
                                    ws.append(linha)
                        else:
                            texto = pagina.extract_text()
                            if texto:
                                for linha in texto.split("\n"):
                                    ws.append([linha])

                output = io.BytesIO()
                wb.save(output)
                dados_excel = output.getvalue()

                st.success("🎉 PDF convertido com sucesso!")

                nome_final = arquivo_pdf.name.replace(".pdf", ".xlsx")
                st.download_button(
                    label="📥 Baixar Arquivo Excel",
                    data=dados_excel,
                    file_name=nome_final,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )

                enviar_para_supabase(f"conversoes/{nome_final}", dados_excel)


# =========================================================================
# MÓDULO: CONCILIAÇÃO BANCÁRIA
# =========================================================================
elif opcao == "📊 Conciliação Bancária":
    st.title("📊 Conciliação de Lançamentos Financeiros")
    st.markdown("Cruza os dados da Planilha de Lançamentos (Braslog) com o Extrato Bancário do Sicoob.")

    col1, col2 = st.columns(2)
    with col1:
        arq_braslog = st.file_uploader("Upload: Balanço Braslog (XLSX)", type=["xlsx"])
    with col2:
        arq_sicoob = st.file_uploader("Upload: Extrato Sicoob (CSV)", type=["csv"])

    if arq_braslog and arq_sicoob:
        if st.button("🔄 Executar Conciliação"):
            with st.spinner("Cruzando dados..."):
                try:
                    # --- Leitura da planilha de lançamentos ---
                    df_plan = pd.read_excel(arq_braslog, sheet_name="3", header=5)
                    df_plan.columns = [str(c).strip() for c in df_plan.columns]
                    df_plan = df_plan[df_plan["CONTA"] == "SIC"].copy()

                    df_plan["DATA"] = pd.to_datetime(df_plan["DATA"], errors="coerce").dt.normalize()
                    df_plan["Recebido"] = pd.to_numeric(df_plan["Recebido"], errors="coerce").fillna(0)
                    df_plan["Pago"] = pd.to_numeric(df_plan["Pago"], errors="coerce").fillna(0)

                    def compoe_valor_direcao(row):
                        if row["Recebido"] > 0:
                            return pd.Series({"Valor": row["Recebido"], "Direcao": "C"})
                        if row["Pago"] > 0:
                            return pd.Series({"Valor": row["Pago"], "Direcao": "D"})
                        return pd.Series({"Valor": 0.0, "Direcao": None})

                    df_plan[["Valor", "Direcao"]] = df_plan.apply(compoe_valor_direcao, axis=1)
                    df_plan = df_plan[df_plan["Valor"] > 0].copy()
                    df_plan["Valor"] = df_plan["Valor"].round(2)

                    # --- Leitura do extrato Sicoob ---
                    df_ext = pd.read_csv(
                        arq_sicoob,
                        sep=",",
                        decimal=",",
                        thousands=".",
                        encoding="utf-8",
                    )
                    df_ext.columns = [str(c).strip() for c in df_ext.columns]
                    df_ext["data"] = pd.to_datetime(df_ext["data"], format="%d/%m/%Y", errors="coerce").dt.normalize()
                    df_ext["valor"] = pd.to_numeric(df_ext["valor"], errors="coerce")
                    df_ext = df_ext.rename(columns={"data": "DATA", "valor": "Valor"})
                    df_ext["Direcao"] = df_ext["tipo"]
                    df_ext["Valor"] = df_ext["Valor"].round(2)

                    # --- Filtro de período ---
                    data_min = df_ext["DATA"].min()
                    data_max = df_ext["DATA"].max()
                    df_plan = df_plan[(df_plan["DATA"] >= data_min) & (df_plan["DATA"] <= data_max)].copy()

                    # --- Numeração de ocorrências repetidas ---
                    df_plan["Seq"] = df_plan.groupby(["DATA", "Valor", "Direcao"]).cumcount()
                    df_ext["Seq"] = df_ext.groupby(["DATA", "Valor", "Direcao"]).cumcount()

                    # --- Merge ---
                    colunas_plan = ["FORNECEDOR", "DESTINO", "Categoria", "SUB Categoria", "OBSERVAÇÃO"]
                    colunas_ext = ["historico", "detalhes"]

                    df_plan_keep = df_plan[["DATA", "Valor", "Direcao", "Seq"] + [c for c in colunas_plan if c in df_plan.columns]].copy()
                    df_ext_keep = df_ext[["DATA", "Valor", "Direcao", "Seq"] + [c for c in colunas_ext if c in df_ext.columns]].copy()

                    df = pd.merge(
                        df_plan_keep,
                        df_ext_keep,
                        on=["DATA", "Valor", "Direcao", "Seq"],
                        how="outer",
                        suffixes=("_plan", "_ext"),
                        indicator=True,
                    )

                    df["Status"] = df["_merge"].map({
                        "both": "Ok",
                        "left_only": "Falta no Extrato",
                        "right_only": "Falta na Planilha",
                    })
                    df = df.drop(columns=["_merge"])
                    df["Direcao"] = df["Direcao"].map({"C": "Entrada", "D": "Saída"})

                    # Mês dinâmico (não fixa em 4/5/6)
                    nomes_meses = {
                        1: "Janeiro", 2: "Fevereiro", 3: "Março", 4: "Abril",
                        5: "Maio", 6: "Junho", 7: "Julho", 8: "Agosto",
                        9: "Setembro", 10: "Outubro", 11: "Novembro", 12: "Dezembro",
                    }
                    df["Mes"] = df["DATA"].dt.month.map(nomes_meses)

                    ordem = [
                        "Status", "Mes", "DATA", "Valor", "Direcao",
                        "FORNECEDOR", "DESTINO", "Categoria", "SUB Categoria", "OBSERVAÇÃO",
                        "historico", "detalhes",
                    ]
                    colunas_finais = [c for c in ordem if c in df.columns]
                    df = df[colunas_finais]
                    df = df.sort_values(["DATA", "Status", "Valor"]).reset_index(drop=True)

                    st.success("Conciliação efetuada com sucesso!")

                    # Resumo por status
                    st.subheader("Resumo")
                    st.dataframe(df["Status"].value_counts().reset_index().rename(columns={"index": "Status", "Status": "Quantidade"}))

                    # Prévia
                    st.subheader("Prévia dos Resultados")
                    st.dataframe(df, use_container_width=True)

                    # Download
                    output = io.BytesIO()
                    with pd.ExcelWriter(output, engine="openpyxl") as writer:
                        df.to_excel(writer, index=False, sheet_name="Resultado")
                    dados_conciliacao = output.getvalue()

                    st.download_button(
                        label="📥 Baixar Resultado da Conciliação",
                        data=dados_conciliacao,
                        file_name="Resultado_Conciliacao.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )

                    enviar_para_supabase("conciliacoes/Resultado_Conciliacao.xlsx", dados_conciliacao)

                except Exception as e:
                    st.error(f"Erro ao processar conciliação: {e}")


# =========================================================================
# MÓDULO: DIFERENÇAS BRASLOG VS MH
# =========================================================================
elif opcao == "💸 Diferenças Braslog vs MH":
    st.title("💸 Análise de Transferências (Braslog <-> MH)")
    st.markdown("Filtra e calcula o saldo das transferências mútuas entre Braslog e MH Gestão.")

    arq_diferenca = st.file_uploader("Upload: Planilha de Lançamentos (XLSX)", type=["xlsx"])

    if arq_diferenca:
        if st.button("🔍 Analisar Diferenças"):
            with st.spinner("Filtrando e calculando saldos..."):
                try:
                    df = pd.read_excel(arq_diferenca, sheet_name="3", header=5)
                    df.columns = [str(c).strip() for c in df.columns]

                    col_obs = [c for c in df.columns if "OBS" in c.upper()][0]
                    df["DATA"] = pd.to_datetime(df["DATA"], errors="coerce").dt.normalize()
                    df["Recebido"] = pd.to_numeric(df["Recebido"], errors="coerce").fillna(0)
                    df["Pago"] = pd.to_numeric(df["Pago"], errors="coerce").fillna(0)

                    colunas_busca = [
                        "DESTINO", "FORNECEDOR", "Categoria", "SUB Categoria",
                        "Centro de custo", col_obs,
                    ]
                    texto = (
                        df[[c for c in colunas_busca if c in df.columns]]
                        .fillna("")
                        .astype(str)
                        .agg(" ".join, axis=1)
                        .str.upper()
                    )

                    masc_mh = texto.str.contains("MH GEST", na=False) | texto.str.contains("MHGEST", na=False)
                    masc_aportes = (df["Categoria"].fillna("").str.upper() == "APORTES") | texto.str.contains("EMPREST", na=False)
                    masc_adiant = (
                        texto.str.contains("ADIANT", na=False)
                        | texto.str.contains("ANTECIP", na=False)
                        | texto.str.contains("SALARI", na=False)
                        | texto.str.contains("SALÁR", na=False)
                    )

                    df_filtrado = df[masc_mh & masc_aportes & ~masc_adiant].copy()

                    def direcao(row):
                        if row["Pago"] > 0:
                            return "Braslog -> MH (MH deve a Braslog)"
                        if row["Recebido"] > 0:
                            return "MH -> Braslog (Braslog deve a MH)"
                        return ""

                    df_filtrado["Direção"] = df_filtrado.apply(direcao, axis=1)

                    total_pago = df_filtrado["Pago"].sum()
                    total_recebido = df_filtrado["Recebido"].sum()
                    saldo = total_pago - total_recebido

                    if saldo > 0.005:
                        devedor, credor, valor = "MH", "Braslog", saldo
                    elif saldo < -0.005:
                        devedor, credor, valor = "Braslog", "MH", abs(saldo)
                    else:
                        devedor, credor, valor = None, None, 0.0

                    ordem = [
                        "Direção", "DATA", "CONTA", "BANCO",
                        "DESTINO", "FORNECEDOR", "Centro de custo", "Categoria", "SUB Categoria",
                        "Recebido", "Pago", col_obs,
                    ]
                    colunas_finais = [c for c in ordem if c in df_filtrado.columns]
                    df_filtrado = df_filtrado[colunas_finais]
                    df_filtrado = df_filtrado.sort_values(["DATA"]).reset_index(drop=True)

                    # Linhas de resumo
                    linhas_resumo = [
                        {c: "" for c in colunas_finais},
                        {**{c: "" for c in colunas_finais}, "Direção": "RESUMO DAS TRANSFERÊNCIAS BRASLOG <-> MH"},
                        {**{c: "" for c in colunas_finais}, "Direção": "Total pago por Braslog a MH (Pago):", "Pago": total_pago},
                        {**{c: "" for c in colunas_finais}, "Direção": "Total recebido por Braslog de MH (Recebido):", "Recebido": total_recebido},
                        {**{c: "" for c in colunas_finais}, "Direção": "Saldo líquido (Pago - Recebido):", "Pago": saldo},
                    ]
                    if devedor:
                        linhas_resumo.append(
                            {**{c: "" for c in colunas_finais}, "Direção": f"=> {devedor} DEVE a {credor}:", "Pago": valor}
                        )
                    df_saida = pd.concat([df_filtrado, pd.DataFrame(linhas_resumo)], ignore_index=True)

                    st.success("Análise concluída!")

                    # Resumo destacado
                    st.subheader("Resumo Financeiro")
                    m1, m2, m3 = st.columns(3)
                    m1.metric("Total Pago (Braslog → MH)", f"R$ {total_pago:,.2f}")
                    m2.metric("Total Recebido (MH → Braslog)", f"R$ {total_recebido:,.2f}")
                    m3.metric("Saldo Líquido", f"R$ {saldo:,.2f}")
                    if devedor:
                        st.info(f"➡️ **{devedor} DEVE a {credor}: R$ {valor:,.2f}**")
                    else:
                        st.info("➡️ Contas zeradas.")

                    st.subheader("Detalhamento das Transferências")
                    st.dataframe(df_saida, use_container_width=True)

                    output = io.BytesIO()
                    with pd.ExcelWriter(output, engine="openpyxl") as writer:
                        df_saida.to_excel(writer, index=False, sheet_name="Diferenças")
                    dados_dif = output.getvalue()

                    st.download_button(
                        label="📥 Baixar Relatório de Diferenças",
                        data=dados_dif,
                        file_name="Diferencas_Abas.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )

                    enviar_para_supabase("diferencas/Diferencas_Abas.xlsx", dados_dif)

                except Exception as e:
                    st.error(f"Erro ao processar análise: {e}")
